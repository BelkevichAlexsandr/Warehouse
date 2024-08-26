import datetime
from io import BytesIO
from typing import Any

import openpyxl
from fastapi import UploadFile, File, HTTPException, status
from sqlalchemy.exc import DBAPIError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import and_

from app.databases.dao.manufacturer import ManufacturerDAO
from app.databases.dao.serial_number import SerialNumberDAO
from app.databases.dao.supplier import SupplierDAO
from app.databases.dao.transaction_layer import TransactionLayer
from app.databases.dao.warehouse import WarehouseDAO
from app.models import SerialNumberStatusEnum, Warehouse
from app.schemas.warehouse import WarehouseModel, PatchWarehouseModel


class WarehouseService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    async def get_warehouse_with_serial_numbers(
        self,
        search: str | None = None,
        need_id: list[int] | None = None,
    ):
        async with WarehouseDAO(self.db) as dao:
            warehouse_ids = []
            if need_id:
                warehouse_ids = [dao.model.id.in_(need_id)]
            return await dao.get_list(
                where=[
                    dao.model.name.ilike(f"%{search}%"),
                    dao.model.deleted_at.is_(None),
                    *warehouse_ids,
                ],
                select_in_load=[dao.model.serial_numbers]
            )

    async def get_warehouse(self, item_id: int):
        async with WarehouseDAO(self.db) as dao:
            return await dao.get_one(
                where=[
                    dao.model.id == item_id,
                    dao.model.deleted_at.is_(None),
                ],
            )

    async def create_warehouse(self, request: WarehouseModel):
        async with WarehouseDAO(self.db) as dao:
            if await dao.get_one(
                where=[
                    dao.model.name == request.name,
                    dao.model.article == request.article,
                    dao.model.deleted_at.is_(None),
                ],
            ):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Товар с такими параметрами уже существует. Создание нового не возможно.",
                )
            return await dao.create_item(request)

    async def update_warehouse(self, item_id: int, request: PatchWarehouseModel):
        async with WarehouseDAO(self.db) as dao:
            if not await dao.get_one(
                where=[
                    dao.model.id == item_id,
                    dao.model.deleted_at.is_(None),
                ],
            ):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Товар с таким id {item_id} не существует. Обновление не возможно.",
                )
            return await dao.update_item(item_id=item_id, item=request)

    async def delete_warehouse(self, item_id: int):
        async with WarehouseDAO(self.db) as dao:
            if not await dao.get_one(
                where=[
                    dao.model.id == item_id,
                    dao.model.deleted_at.is_(None),
                ],
            ):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Товар с таким id {item_id} не существует. Удаление не возможно.",
                )
            await dao.update_item(
                item_id=item_id,
                item={"deleted_at": datetime.datetime.now()},
            )

    async def parse_excel_file(
        self,
        file: UploadFile = File(...),
    ) -> None:
        tables: dict[str, list[dict[str, Any]]] = self._parse_file(file=file)

        if not tables.get('ЗАКАЗ'):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='Sheet ЗАКАЗ not found',
            )

        new_warehouses: list[dict[str, Any]] = []
        new_serial_numbers = {}
        check_product_warehouse: set = set()
        check_product_serial_number: set = set()

        async with TransactionLayer(self.db):
            warehouse_names = {str(row['Наименование ']) for row in tables.get('ЗАКАЗ') if row['Наименование ']}
            warehouse_articles = {str(row['Артикул']) for row in tables.get('ЗАКАЗ') if row['Артикул']}
            warehouse_records = await WarehouseDAO(self.db).get_warehouse_by_name_and_article_with_serial_number(
                names=list(warehouse_names),
                articles=list(warehouse_articles),
            )
            suppliers = await SupplierDAO(self.db).get_list()
            suppliers = {supplier.name: supplier.id for supplier in suppliers}
            manufacturers = await ManufacturerDAO(self.db).get_list()
            manufacturers = {manufacturer.name: manufacturer.id for manufacturer in manufacturers}
            if warehouse_records:
                check_product_warehouse = {
                    warehouse.name for warehouse in warehouse_records
                }
                check_product_serial_number = {
                    serial_number.name for warehouse in warehouse_records
                    for serial_number in warehouse.serial_numbers
                }

            for row in tables.get('ЗАКАЗ'):
                if not row.get('Артикул'):
                    continue
                if not row.get('Поставщик'):
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f'Не указан поставщик в строке {row}',
                    )
                if not row.get('Производитель'):
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f'Не указан производитель в строке {row}',
                    )
                if row.get('Наименование ') not in check_product_warehouse:
                    warehouse = {
                        'manufacturer_id': manufacturers.get(row.get('Производитель')),
                        'supplier_id': suppliers.get(row.get('Поставщик')),
                        'article': str(row.get('Артикул')),
                        'name': str(row.get('Наименование ')),
                        'warranty': row.get('Гарантия, мес.'),
                        'product_count_in_stock': row.get('кол-во по позиции'),
                    }
                    check_product_warehouse.add(row.get('Наименование '))
                    new_warehouses.append(warehouse)
                if row.get('Серийный номер\nS/N') in check_product_serial_number:
                    continue
                serial_number = {
                        'name': row.get('Серийный номер\nS/N'),
                        'status': SerialNumberStatusEnum.WAREHOUSE,
                        'price_input': row.get('Цена входа'),
                }
                check_product_serial_number.add(row.get('Серийный номер\nS/N'))
                if new_serial_numbers.get(row.get('Наименование ')):
                    new_serial_numbers.get(row.get('Наименование ')).append(serial_number)
                    continue

                new_serial_numbers[row.get('Наименование ')] = [serial_number]

            if new_warehouses:
                try:
                    await WarehouseDAO(self.db).insert_bulk(items=new_warehouses)
                except DBAPIError:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail='DBAPIError warehouse',
                    )
            warehouse_records = await WarehouseDAO(self.db).get_list(
                where=[
                    and_(
                        Warehouse.name.in_(warehouse_names),
                        Warehouse.article.in_(warehouse_articles)
                    )
                ]

            )
            warehouse_records = {
                warehouse.name: warehouse.id for warehouse in warehouse_records
            }
            new_serial_numbers = [
                {'warehouse_id': warehouse_records.get(warehouse_name), **serial}
                for warehouse_name, serials in new_serial_numbers.items()
                for serial in serials
            ]
            if new_serial_numbers:
                try:
                    await SerialNumberDAO(self.db).insert_bulk(items=new_serial_numbers)
                except DBAPIError:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail='DBAPIError serial_numbers',
                    )
            warehouse_records = await WarehouseDAO(self.db).get_warehouse_by_name_with_serial_number_to_stock(
                names=list(check_product_warehouse),
            )
            update_warehouse = [
                {'id': warehouse.id, 'product_count_in_stock': len(warehouse.serial_numbers)}
                for warehouse in warehouse_records
            ]
            if update_warehouse:
                try:
                    await WarehouseDAO(self.db).update_bulk(items=update_warehouse)
                except DBAPIError:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail='DBAPIError warehouse',
                    )

    @staticmethod
    def _parse_file(file: UploadFile = File(...)) -> dict[str, list[dict[str, Any]]]:
        tables: dict[str, list[dict[str, Any]]] = {}
        with BytesIO(file.file.read()) as buffer:
            dataframe = openpyxl.load_workbook(buffer)

        for sheet_name in dataframe.sheetnames:
            table = []
            sheet = dataframe[sheet_name]
            headers_row = [value.value for value in sheet[1]]
            for row in sheet.iter_rows(values_only=True):
                element = {}
                if headers_row[0] == row[0]:
                    continue
                for key, value in zip(headers_row, row):
                    element.update({key: value})
                table.append(element)

            tables.update({sheet_name: table})
        return tables
